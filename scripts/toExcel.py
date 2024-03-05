import openpyxl


def generar_datos_excel(nombre_archivo, datos):
    # Crear un nuevo libro de Excel
    libro = openpyxl.Workbook()

    # Crear una nueva hoja en el libro
    hoja = libro.active
    hoja.title = "Datos"

    # Agregar encabezados
    hoja["A1"] = "ID"
    hoja["B1"] = "TEXT"
    hoja["C1"] = "COMPARTIDO"

    # Iterar sobre los datos y escribir en la hoja de Excel
    for fila, dato in enumerate(datos, start=2):
        hoja.cell(row=fila, column=1, value=dato["ID"])
        hoja.cell(row=fila, column=2, value=dato["TEXT"])
        hoja.cell(row=fila, column=3, value="PEPE" if fila in (2, 3) else "")

    # Combinar celdas en la columna B para los índices 2 y 3
    hoja.merge_cells("B2:B3")

    # Crear una nueva hoja en el libro
    nueva_hoja = libro.create_sheet(title="NuevaHoja")

    # Agregar encabezados en la nueva hoja
    nueva_hoja["A1"] = "ID"
    nueva_hoja["B1"] = "NUEVO_TEXTO"

    # Iterar sobre los datos y escribir en la nueva hoja de Excel
    for fila, dato in enumerate(datos, start=2):
        nueva_hoja.cell(row=fila, column=1, value=dato["ID"])
        nueva_hoja.cell(row=fila, column=2, value=f"NuevoTexto-{fila}")

    # Guardar el libro en un archivo
    libro.save(nombre_archivo)


if __name__ == "__main__":
    # Datos de ejemplo
    datos_ejemplo = [
        {"ID": 1, "TEXT": "Texto1"},
        {"ID": 2, "TEXT": "Texto2"},
        {"ID": 3, "TEXT": "Texto3"},
        # Agrega más datos según sea necesario
    ]

    # Nombre del archivo de salida
    nombre_archivo = "datos_excel_con_hoja_nueva.xlsx"

    # Generar el archivo Excel
    generar_datos_excel(nombre_archivo, datos_ejemplo)

    print(f"Se ha generado el archivo Excel: {nombre_archivo}")
